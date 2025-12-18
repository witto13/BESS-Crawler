"""
Excel export using pandas/openpyxl with formatting.
"""
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_procedures(rows: List[Dict], path: str, sheet_name: str = "Procedures") -> None:
    """
    Write procedures to Excel with formatting.
    """
    if not rows:
        df = pd.DataFrame()
    else:
        df = pd.DataFrame(rows)
    
    # Write to Excel
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format the sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def export_from_db(db_dsn: str, output_path: str, filter_high_confidence: bool = False) -> None:
    """
    Export procedures from database to Excel.
    """
    from psycopg import connect
    import pandas as pd
    
    query = """
    SELECT 
        p.procedure_id,
        p.title_raw,
        p.title_norm,
        p.instrument,
        p.status,
        p.state,
        p.county,
        p.municipality,
        p.municipality_key,
        p.bess_score,
        p.grid_score,
        p.confidence,
        p.developer_company,
        p.capacity_mw,
        p.capacity_mwh,
        p.area_hectares,
        p.decision_date,
        p.created_at,
        p.updated_at,
        COUNT(DISTINCT s.source_id) as source_count,
        COUNT(DISTINCT d.document_id) as document_count
    FROM procedures p
    LEFT JOIN sources s ON p.procedure_id = s.procedure_id
    LEFT JOIN documents d ON s.source_id = d.source_id
    WHERE p.procedure_id != 'test-proc-999'
    """
    
    if filter_high_confidence:
        query += " AND p.confidence = 'high'"
    
    query += " GROUP BY p.procedure_id ORDER BY p.bess_score DESC, p.grid_score DESC"
    
    with connect(db_dsn) as conn:
        df = pd.read_sql_query(query, conn)
    
    # Convert timezone-aware datetimes to timezone-naive for Excel
    datetime_cols = df.select_dtypes(include=['datetime64[ns, UTC]', 'datetime64[ns]']).columns
    for col in datetime_cols:
        if df[col].dtype.tz is not None:
            df[col] = df[col].dt.tz_localize(None)
    
    # Export to Excel with multiple sheets
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # All procedures
        df.to_excel(writer, sheet_name="All Procedures", index=False)
        
        # High confidence only
        if not filter_high_confidence:
            df_high = df[df["confidence"] == "high"].copy()
            df_high.to_excel(writer, sheet_name="High Confidence", index=False)
        
        # Summary statistics
        summary_data = {
            "Metric": [
                "Total Procedures",
                "High Confidence",
                "Medium Confidence",
                "Low Confidence",
                "With BESS Score > 0",
                "With Grid Score > 0",
                "With Capacity (MW)",
                "With Area (Hectares)",
                "With Decision Date",
                "With Company",
                "Avg BESS Score",
                "Avg Grid Score",
                "Total Capacity (MW)",
                "Total Area (Hectares)",
            ],
            "Value": [
                len(df),
                len(df[df["confidence"] == "high"]),
                len(df[df["confidence"] == "medium"]),
                len(df[df["confidence"] == "low"]),
                len(df[df["bess_score"] > 0]),
                len(df[df["grid_score"] > 0]),
                len(df[df["capacity_mw"].notna()]),
                len(df[df["area_hectares"].notna()]),
                len(df[df["decision_date"].notna()]),
                len(df[df["developer_company"].notna()]),
                df["bess_score"].mean(),
                df["grid_score"].mean(),
                df["capacity_mw"].sum() if "capacity_mw" in df.columns else 0,
                df["area_hectares"].sum() if "area_hectares" in df.columns else 0,
            ],
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
    
    # Format all sheets
    workbook = load_workbook(output_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output_path)
    print(f"Exported {len(df)} procedures to {output_path}")


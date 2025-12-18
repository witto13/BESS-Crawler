#!/usr/bin/env python3
"""
Export project entities to Excel with projects, timeline, and diagnostics sheets.
"""
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from apps.db.client import get_pool
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def export_projects_to_excel(output_path: str):
    """
    Export project entities to Excel with multiple sheets.
    """
    pool = get_pool()
    
    with pool.connection() as conn:
        # Projects sheet (all projects)
        projects_query = """
        SELECT 
            pe.project_id,
            pe.state,
            pe.municipality_name,
            pe.county,
            pe.canonical_project_name,
            pe.maturity_stage,
            pe.legal_basis_best,
            pe.project_components,
            pe.developer_company_best,
            pe.site_location_best,
            pe.capacity_mw_best,
            pe.capacity_mwh_best,
            pe.area_hectares_best,
            pe.first_seen_date,
            pe.last_seen_date,
            pe.max_confidence,
            pe.needs_review,
            COUNT(DISTINCT pp.procedure_id) as number_of_procedures,
            COUNT(DISTINCT CASE WHEN p.procedure_type != 'UNKNOWN' THEN p.procedure_id END) as number_of_known_procedures,
            COUNT(DISTINCT s.source_id) as number_of_sources,
            COUNT(DISTINCT d.document_id) as number_of_documents
        FROM project_entities pe
        LEFT JOIN project_procedures pp ON pp.project_id = pe.project_id
        LEFT JOIN procedures p ON p.procedure_id = pp.procedure_id
        LEFT JOIN sources s ON s.procedure_id = p.procedure_id
        LEFT JOIN documents d ON d.source_id = s.source_id
        WHERE pe.state = 'BB'
        GROUP BY pe.project_id
        ORDER BY pe.first_seen_date DESC, pe.maturity_stage
        """
        
        # High-confidence projects sheet
        high_confidence_query = """
        SELECT 
            pe.project_id,
            pe.state,
            pe.municipality_name,
            pe.county,
            pe.canonical_project_name,
            pe.maturity_stage,
            pe.legal_basis_best,
            pe.project_components,
            pe.developer_company_best,
            pe.site_location_best,
            pe.capacity_mw_best,
            pe.capacity_mwh_best,
            pe.area_hectares_best,
            pe.first_seen_date,
            pe.last_seen_date,
            pe.max_confidence,
            pe.needs_review,
            COUNT(DISTINCT pp.procedure_id) as number_of_procedures,
            COUNT(DISTINCT CASE WHEN p.procedure_type != 'UNKNOWN' THEN p.procedure_id END) as number_of_known_procedures,
            COUNT(DISTINCT s.source_id) as number_of_sources,
            COUNT(DISTINCT d.document_id) as number_of_documents
        FROM project_entities pe
        LEFT JOIN project_procedures pp ON pp.project_id = pe.project_id
        LEFT JOIN procedures p ON p.procedure_id = pp.procedure_id
        LEFT JOIN sources s ON s.procedure_id = p.procedure_id
        LEFT JOIN documents d ON d.source_id = s.source_id
        WHERE pe.state = 'BB'
        AND pe.max_confidence >= 0.6
        AND EXISTS (
            SELECT 1 FROM project_procedures pp2
            JOIN procedures p2 ON p2.procedure_id = pp2.procedure_id
            WHERE pp2.project_id = pe.project_id
            AND p2.procedure_type IS NOT NULL
            AND p2.procedure_type != 'UNKNOWN'
        )
        GROUP BY pe.project_id
        ORDER BY pe.max_confidence DESC, pe.first_seen_date DESC
        """
        
        df_projects = pd.read_sql_query(projects_query, conn)
        df_high_confidence = pd.read_sql_query(high_confidence_query, conn)
        
        # Project timeline sheet
        timeline_query = """
        SELECT 
            pp.project_id,
            p.procedure_id,
            p.procedure_type,
            p.decision_date,
            s.discovery_source,
            s.discovery_path,
            p.title_raw,
            p.evidence_snippets
        FROM project_procedures pp
        JOIN procedures p ON p.procedure_id = pp.procedure_id
        LEFT JOIN sources s ON s.procedure_id = p.procedure_id
        WHERE p.state = 'BB'
        ORDER BY pp.project_id, p.decision_date, p.created_at
        """
        
        df_timeline = pd.read_sql_query(timeline_query, conn)
        
        # Extract first evidence snippet for timeline
        if 'evidence_snippets' in df_timeline.columns:
            import json
            def get_first_snippet(ev):
                if not ev:
                    return None
                try:
                    snippets = json.loads(ev) if isinstance(ev, str) else ev
                    if isinstance(snippets, list) and snippets:
                        return snippets[0][:250] if isinstance(snippets[0], str) else str(snippets[0])[:250]
                except:
                    pass
                return None
            df_timeline['top_evidence_snippet'] = df_timeline['evidence_snippets'].apply(get_first_snippet)
            df_timeline = df_timeline.drop(columns=['evidence_snippets'])
        
        # Diagnostics sheet
        diagnostics = {}
        
        # Procedure counts
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM procedures WHERE procedure_id != 'test-proc-999' AND state = 'BB'")
        diagnostics['procedures_total'] = cur.fetchone()[0]
        
        # Skipped procedures (sources without procedures)
        cur.execute("SELECT COUNT(*) FROM sources WHERE procedure_id IS NULL")
        diagnostics['procedures_skipped_container'] = cur.fetchone()[0]
        
        # Valid procedures
        cur.execute("SELECT COUNT(*) FROM procedures WHERE procedure_id != 'test-proc-999' AND state = 'BB' AND procedure_type IS NOT NULL")
        diagnostics['valid_procedures'] = cur.fetchone()[0]
        
        # Projects
        cur.execute("SELECT COUNT(*) FROM project_entities WHERE state = 'BB'")
        diagnostics['projects_total'] = cur.fetchone()[0]
        
        # Projects by maturity
        cur.execute("""
            SELECT maturity_stage, COUNT(*) 
            FROM project_entities 
            WHERE state = 'BB' 
            GROUP BY maturity_stage 
            ORDER BY COUNT(*) DESC
        """)
        maturity_counts = dict(cur.fetchall())
        for stage, count in maturity_counts.items():
            diagnostics[f'projects_{stage}'] = count
        
        # Projects by county
        cur.execute("""
            SELECT county, COUNT(*) 
            FROM project_entities 
            WHERE state = 'BB' 
            GROUP BY county 
            ORDER BY COUNT(*) DESC
            LIMIT 10
        """)
        county_counts = dict(cur.fetchall())
        for county, count in county_counts.items():
            diagnostics[f'projects_county_{county}'] = count
        
        # Discovery source breakdown
        cur.execute("""
            SELECT COALESCE(s.discovery_source, 'UNKNOWN'), COUNT(DISTINCT p.procedure_id)
            FROM procedures p
            LEFT JOIN sources s ON s.procedure_id = p.procedure_id
            WHERE p.procedure_id != 'test-proc-999' AND p.state = 'BB'
            GROUP BY s.discovery_source
        """)
        source_counts = dict(cur.fetchall())
        for source, count in source_counts.items():
            diagnostics[f'source_{source}'] = count
        
        df_diagnostics = pd.DataFrame([
            {"Metric": k, "Value": v} for k, v in diagnostics.items()
        ])
    
    # Convert timezone-aware datetimes
    datetime_cols = df_projects.select_dtypes(include=['datetime64[ns, UTC]', 'datetime64[ns]']).columns
    for col in datetime_cols:
        if df_projects[col].dtype.tz is not None:
            df_projects[col] = df_projects[col].dt.tz_localize(None)
    
    datetime_cols = df_timeline.select_dtypes(include=['datetime64[ns, UTC]', 'datetime64[ns]']).columns
    for col in datetime_cols:
        if df_timeline[col].dtype.tz is not None:
            df_timeline[col] = df_timeline[col].dt.tz_localize(None)
    
    # Write to Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_projects.to_excel(writer, sheet_name="all_projects", index=False)
        df_high_confidence.to_excel(writer, sheet_name="high_confidence_projects", index=False)
        df_timeline.to_excel(writer, sheet_name="project_timeline", index=False)
        df_diagnostics.to_excel(writer, sheet_name="diagnostics", index=False)
        
        # Format sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Exported {len(df_projects)} projects to {output_path}")


if __name__ == "__main__":
    output_path = sys.argv[1] if len(sys.argv) > 1 else "exports/bess_projects.xlsx"
    export_projects_to_excel(output_path)

